import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from doc_gen import generate_payment_advice, generate_summary_pdf
from config import Config

class BookkeepingSystem:
    def __init__(self):
        # Default active sheet is based on TODAY
        self.active_sheet_name = self.get_sheet_name_for_date(datetime.now())
        self.ensure_file_exists()

    def get_sheet_name_for_date(self, date_obj):
        """
        Determines the correct sheet name based on the specific transaction date.
        Format: Transactions_YYYY_YY (e.g., Transactions_2026_27)
        """
        if date_obj.month >= 4: # April to Dec
            start_year = date_obj.year
            end_year = date_obj.year + 1
        else: # Jan to March
            start_year = date_obj.year - 1
            end_year = date_obj.year
            
        short_end = str(end_year)[-2:]
        return f"{Config.TXN_PREFIX}{start_year}_{short_end}"

    def ensure_file_exists(self):
        if not os.path.exists(Config.DB_FILENAME):
            wb = openpyxl.Workbook()
            ws_limits = wb.active
            ws_limits.title = Config.SHEET_LIMITS
            # Defines the column for Opening Balance / Previous Balance
            ws_limits.append(["Department", "Previous_balance"]) 
            
            wb.create_sheet(self.active_sheet_name)
            wb.save(Config.DB_FILENAME)
        else:
            # Check if active sheet exists
            try:
                wb = openpyxl.load_workbook(Config.DB_FILENAME)
                if self.active_sheet_name not in wb.sheetnames:
                    wb.create_sheet(self.active_sheet_name)
                    wb.save(Config.DB_FILENAME)
            except: pass

    def _ensure_fy_sheet_exists(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            return wb.create_sheet(sheet_name)
        return wb[sheet_name]

    def get_subsidiaries(self):
        try:
            wb = openpyxl.load_workbook(Config.DB_FILENAME, data_only=True)
            ws = wb[Config.SHEET_LIMITS]
            subs = []
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if val: subs.append(val)
            return subs
        except: return []

    def get_limit_info(self, subsidiary):
        try:
            wb = openpyxl.load_workbook(Config.DB_FILENAME, data_only=True)
            ws = wb[Config.SHEET_LIMITS]
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == subsidiary:
                    val = ws.cell(row=row, column=2).value
                    return int(val) if val else 0
            return 0
        except: return 0

    def _get_or_create_subsidiary_columns(self, wb, ws, subsidiary):
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col in range(1, ws.max_column + 2, 3):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value == subsidiary:
                return col
            if cell_value is None:
                start_col = col
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+2)
                title_cell = ws.cell(row=1, column=start_col, value=subsidiary)
                title_cell.font = bold_font
                title_cell.alignment = center_align
                for c in range(start_col, start_col+3):
                    ws.cell(row=1, column=c).border = thin_border

                headers = ["PPA_Number", "Date", "Amount"]
                for i, header in enumerate(headers):
                    cell = ws.cell(row=2, column=start_col + i, value=header)
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.border = thin_border
                return start_col
        return 1

    def _fmt_money(self, value):
        try: value = int(value)
        except: return str(value)
        s = str(value)
        if len(s) <= 3: return f"₹ {s}"
        last_three = s[-3:]
        remaining = s[:-3]
        formatted_remaining = ""
        for i, digit in enumerate(reversed(remaining)):
            if i > 0 and i % 2 == 0: formatted_remaining = "," + formatted_remaining
            formatted_remaining = digit + formatted_remaining
        return f"₹ {formatted_remaining},{last_three}"

    def save_batch(self, subsidiary, batch_list):
        try:
            wb = openpyxl.load_workbook(Config.DB_FILENAME)
            first_date = batch_list[0][1]
            target_sheet_name = self.get_sheet_name_for_date(first_date)
            ws = self._ensure_fy_sheet_exists(wb, target_sheet_name)
        except Exception as e: return False, f"Error: {e}"

        limit = self.get_limit_info(subsidiary)
        start_col = self._get_or_create_subsidiary_columns(wb, ws, subsidiary)
        col_ppa = start_col
        col_amt = start_col + 2
        
        current_spent = 0
        existing_ppas = []
        for row in range(3, ws.max_row + 1):
            val_ppa = ws.cell(row=row, column=col_ppa).value
            val_amt = ws.cell(row=row, column=col_amt).value
            if val_ppa: existing_ppas.append(str(val_ppa))
            if isinstance(val_amt, (int, float)): current_spent += val_amt

        batch_total = 0
        new_ppas = []
        for (ppa, _, amt) in batch_list:
            ppa = str(ppa)
            if ppa in existing_ppas: return False, f"Error: PPA {ppa} exists in {target_sheet_name}."
            if ppa in new_ppas: return False, f"Error: PPA {ppa} duplicated."
            new_ppas.append(ppa)
            batch_total += amt
            
        final_total = current_spent + batch_total
        
        if final_total > limit:
            remaining = limit - current_spent
            msg = (f"Limit Exceeded\nApproved: {self._fmt_money(limit)}\nSpent (FY): {self._fmt_money(current_spent)}\n"
                   f"Batch: {self._fmt_money(batch_total)}\nAvailable: {self._fmt_money(remaining)}")
            return False, msg

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        current_row = 3
        while ws.cell(row=current_row, column=col_ppa).value is not None:
            current_row += 1
            
        for (ppa, date_obj, amt) in batch_list:
            ws.cell(row=current_row, column=col_ppa, value=ppa).border = thin_border
            d_cell = ws.cell(row=current_row, column=col_ppa+1, value=date_obj)
            d_cell.number_format = 'DD-MM-YYYY'
            d_cell.border = thin_border
            a_cell = ws.cell(row=current_row, column=col_ppa+2, value=amt)
            a_cell.number_format = '"₹" #,##0'
            a_cell.border = thin_border
            current_row += 1

        try: wb.save(Config.DB_FILENAME)
        except PermissionError: return False, "Error: File open."
        return True, f"Saved to {target_sheet_name}."

    def save_allocation_batch(self, subsidiary, batch_list):
        try:
            wb = openpyxl.load_workbook(Config.DB_FILENAME)
            ws = wb[Config.SHEET_LIMITS]
        except Exception as e: return False, str(e)

        target_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == subsidiary:
                target_row = r
                break
        
        if not target_row:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1, value=subsidiary)
            ws.cell(row=target_row, column=2, value=0)

        curr_limit_cell = ws.cell(row=target_row, column=2)
        current_limit = curr_limit_cell.value if curr_limit_cell.value else 0
        
        current_col = 3
        while ws.cell(row=target_row, column=current_col).value is not None:
            current_col += 2 
            
        total_added = 0
        for (_, date_obj, amt) in batch_list:
            total_added += amt
            alloc_num = int((current_col - 1) / 2)
            
            # Suffix logic
            if 11 <= (alloc_num % 100) <= 13: suffix = "th"
            else:
                rem = alloc_num % 10
                suffix = {1:"st", 2:"nd", 3:"rd"}.get(rem, "th")
            
            header_title = f"{alloc_num}{suffix} allocation"
            header_date = f"Date_{alloc_num}"
            
            if ws.cell(row=1, column=current_col).value is None:
                ws.cell(row=1, column=current_col, value=header_title).font = Font(bold=True)
                ws.cell(row=1, column=current_col+1, value=header_date).font = Font(bold=True)

            ws.cell(row=target_row, column=current_col, value=amt)
            d_cell = ws.cell(row=target_row, column=current_col+1, value=date_obj)
            d_cell.number_format = 'DD-MM-YYYY'
            current_col += 2

        # Update Column 2 to reflect total accumulated limit
        curr_limit_cell.value = current_limit + total_added

        try: wb.save(Config.DB_FILENAME)
        except PermissionError: return False, "File open."
        return True, f"Allocated {self._fmt_money(total_added)}."

    def get_summary_report(self):
        try:
            df_limits = pd.read_excel(Config.DB_FILENAME, sheet_name=Config.SHEET_LIMITS)
            active_sheet = self.get_sheet_name_for_date(datetime.now())
            wb = openpyxl.load_workbook(Config.DB_FILENAME, data_only=True) 
            if active_sheet in wb.sheetnames:
                ws = wb[active_sheet]
            else:
                return []
        except: return []

        sub_col_map = {}
        for col in range(1, ws.max_column + 2):
            val = ws.cell(row=1, column=col).value
            if val: sub_col_map[val] = col

        summary_data = []
        for index, row in df_limits.iterrows():
            sub_name = row["Department"] if "Department" in row else row["Subsidiary"]
            
            if "Previous_balance" in row:
                limit_val = row["Previous_balance"]
            elif "Approved_Limit" in row:
                limit_val = row["Approved_Limit"]
            else:
                limit_val = 0
            
            limit = int(limit_val) if pd.notna(limit_val) else 0
            
            q1, q2, q3, q4, total_spent = 0, 0, 0, 0, 0
            if sub_name in sub_col_map:
                start_col = sub_col_map[sub_name]
                col_date = start_col + 1
                col_amt = start_col + 2
                for r in range(3, ws.max_row + 1):
                    amt = ws.cell(row=r, column=col_amt).value
                    dt = ws.cell(row=r, column=col_date).value
                    if isinstance(amt, (int, float)) and isinstance(dt, datetime):
                        total_spent += amt
                        m = dt.month
                        if 4 <= m <= 6: q1 += amt
                        elif 7 <= m <= 9: q2 += amt
                        elif 10 <= m <= 12: q3 += amt
                        elif 1 <= m <= 3: q4 += amt
            remaining = limit - total_spent
            summary_data.append((sub_name, limit, q1, q2, q3, q4, total_spent, remaining))
        return summary_data

    # --- UPDATED: DETAILED QUARTERLY PDF DATA ---
    def get_detailed_report_data(self):
        """
        Calculates Net Opening Balance by stripping current FY allocations from the Total Limit.
        Net Opening = (Col 2 Limit - Current FY Allocations) - Historical Expenditures
        """
        try:
            wb = openpyxl.load_workbook(Config.DB_FILENAME, data_only=True)
            ws_limits = wb[Config.SHEET_LIMITS]
        except: return []
        
        # 1. Determine Financial Year Start
        now = datetime.now()
        start_year = now.year if now.month >= 4 else now.year - 1
        fy_start = datetime(start_year, 4, 1)

        # 2. Scan ALL Transaction Sheets for Expenditures
        historical_spent_map = {} 
        current_fy_exp_map = {}   

        for sheet_name in wb.sheetnames:
            if sheet_name.startswith(Config.TXN_PREFIX):
                ws_txn = wb[sheet_name]
                
                dept_cols = {}
                for col in range(1, ws_txn.max_column + 1):
                    val = ws_txn.cell(row=1, column=col).value
                    if val: dept_cols[col] = val
                
                for col, dept_name in dept_cols.items():
                    if dept_name not in historical_spent_map: historical_spent_map[dept_name] = 0
                    if dept_name not in current_fy_exp_map: current_fy_exp_map[dept_name] = {'q1':0, 'q2':0, 'q3':0, 'q4':0}
                    
                    col_date = col + 1
                    col_amt = col + 2
                    
                    for r in range(3, ws_txn.max_row + 1):
                        d_val = ws_txn.cell(row=r, column=col_date).value
                        a_val = ws_txn.cell(row=r, column=col_amt).value
                        
                        if isinstance(d_val, datetime) and isinstance(a_val, (int, float)):
                            if d_val < fy_start:
                                historical_spent_map[dept_name] += a_val
                            else:
                                m = d_val.month
                                if 4 <= m <= 6: current_fy_exp_map[dept_name]['q1'] += a_val
                                elif 7 <= m <= 9: current_fy_exp_map[dept_name]['q2'] += a_val
                                elif 10 <= m <= 12: current_fy_exp_map[dept_name]['q3'] += a_val
                                elif 1 <= m <= 3: current_fy_exp_map[dept_name]['q4'] += a_val

        # 3. Process Allocations & Calculate Final Balances
        detailed_data = []
        
        for r in range(2, ws_limits.max_row + 1):
            sub_name = ws_limits.cell(row=r, column=1).value
            if not sub_name: continue
            
            # A. Grand Total Limit (From Column 2)
            # This includes Opening + ALL Allocations made to date
            col2_val = ws_limits.cell(row=r, column=2).value
            grand_total_limit = int(col2_val) if isinstance(col2_val, (int, float)) else 0
            
            # B. Identify "Current Year" Allocations
            current_fy_alloc_sum = 0
            q_alloc = {'q1':0, 'q2':0, 'q3':0, 'q4':0}
            
            curr_col = 3
            while curr_col <= ws_limits.max_column:
                amt = ws_limits.cell(row=r, column=curr_col).value
                dt = ws_limits.cell(row=r, column=curr_col+1).value
                
                if isinstance(amt, (int, float)):
                    if isinstance(dt, datetime):
                        if dt >= fy_start:
                            # This is a Current FY Allocation
                            current_fy_alloc_sum += amt
                            m = dt.month
                            if 4 <= m <= 6: q_alloc['q1'] += amt
                            elif 7 <= m <= 9: q_alloc['q2'] += amt
                            elif 10 <= m <= 12: q_alloc['q3'] += amt
                            elif 1 <= m <= 3: q_alloc['q4'] += amt
                    # Note: No date or old date implies historical, which is fine.
                    # We only need to subtract Current FY to get Opening Limit.
                curr_col += 2

            # C. Calculate Net Opening Balance
            # Opening Limit = Grand Total - Allocations this year
            opening_limit = grand_total_limit - current_fy_alloc_sum
            
            # Net Opening Balance = Opening Limit - Historical Expenditures
            past_spent = historical_spent_map.get(sub_name, 0)
            net_opening_balance = opening_limit - past_spent

            # D. Calculate Current FY Running Balances
            my_exp = current_fy_exp_map.get(sub_name, {'q1':0, 'q2':0, 'q3':0, 'q4':0})
            
            # Q1
            q1_add = q_alloc['q1']
            q1_exp = my_exp['q1']
            q1_bal = (net_opening_balance + q1_add) - q1_exp
            
            # Q2
            q2_add = q_alloc['q2']
            q2_exp = my_exp['q2']
            q2_bal = (q1_bal + q2_add) - q2_exp
            
            # Q3
            q3_add = q_alloc['q3']
            q3_exp = my_exp['q3']
            q3_bal = (q2_bal + q3_add) - q3_exp
            
            # Q4
            q4_add = q_alloc['q4']
            q4_exp = my_exp['q4']
            q4_bal = (q3_bal + q4_add) - q4_exp
            
            row_tuple = (
                sub_name, net_opening_balance,
                q1_add, q1_exp, q1_bal,
                q2_add, q2_exp, q2_bal,
                q3_add, q3_exp, q3_bal,
                q4_add, q4_exp, q4_bal
            )
            detailed_data.append(row_tuple)
            
        return detailed_data

    def create_word_advice(self, subsidiary, date_str, transaction_list):
        return generate_payment_advice(subsidiary, date_str, transaction_list)

    def create_dashboard_pdf(self, summary_data):
        return generate_summary_pdf(summary_data)

    # --- UNIFIED LEDGER SEARCH ---
    def search_transactions(self, subsidiary=None, ppa_text=None, quarter=None):
        results = []
        try:
            wb = openpyxl.load_workbook(Config.DB_FILENAME, data_only=True)
            ws_limits = wb[Config.SHEET_LIMITS]
            for row in range(2, ws_limits.max_row + 1):
                sub_name = ws_limits.cell(row=row, column=1).value
                if not sub_name: continue
                if subsidiary and subsidiary != "All Departments" and sub_name != subsidiary:
                    continue
                current_col = 3
                while current_col <= ws_limits.max_column:
                    amt = ws_limits.cell(row=row, column=current_col).value
                    date_val = ws_limits.cell(row=row, column=current_col+1).value
                    if isinstance(amt, (int, float)) and isinstance(date_val, datetime):
                        if ppa_text:
                            if "ALLOC" not in ppa_text.upper(): pass 
                        if quarter and quarter != "All":
                            m = date_val.month
                            row_q = ""
                            if 4 <= m <= 6: row_q = "Q1"
                            elif 7 <= m <= 9: row_q = "Q2"
                            elif 10 <= m <= 12: row_q = "Q3"
                            elif 1 <= m <= 3: row_q = "Q4"
                            if row_q != quarter: 
                                current_col += 2
                                continue
                        alloc_num = int((current_col - 1) / 2)
                        ref_text = f"Allocation ({alloc_num})"
                        results.append({"sub": sub_name, "ref": ref_text, "date": date_val, "amt": amt, "type": "ALLOC"})
                    current_col += 2
        except: pass

        try:
            active_sheet = self.get_sheet_name_for_date(datetime.now())
            if active_sheet in wb.sheetnames:
                ws_txn = wb[active_sheet]
                for col in range(1, ws_txn.max_column + 1, 3):
                    sub_name = ws_txn.cell(row=1, column=col).value
                    if not sub_name: continue
                    if subsidiary and subsidiary != "All Departments" and sub_name != subsidiary: continue
                    for row in range(3, ws_txn.max_row + 1):
                        ppa = ws_txn.cell(row=row, column=col).value
                        date_val = ws_txn.cell(row=row, column=col+1).value
                        amt = ws_txn.cell(row=row, column=col+2).value
                        if not ppa: continue 
                        if ppa_text and str(ppa_text).upper() not in str(ppa).upper(): continue
                        if quarter and quarter != "All":
                            if not isinstance(date_val, datetime): continue
                            m = date_val.month
                            row_q = ""
                            if 4 <= m <= 6: row_q = "Q1"
                            elif 7 <= m <= 9: row_q = "Q2"
                            elif 10 <= m <= 12: row_q = "Q3"
                            elif 1 <= m <= 3: row_q = "Q4"
                            if row_q != quarter: continue
                        results.append({"sub": sub_name, "ref": str(ppa), "date": date_val, "amt": amt, "type": "PPA"})
        except: pass
        results.sort(key=lambda x: x["date"], reverse=True)
        final_output = []
        for item in results:
            final_output.append((item["sub"], item["ref"], item["date"], item["amt"]))
        return final_output